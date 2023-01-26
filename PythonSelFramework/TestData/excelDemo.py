import openpyxl

book = openpyxl.load_workbook(
    "C:\\Users\\Alex\\PycharmProjects\\pythonAQAdraft\\PythonSelFramework\\TestData\\DemoFile.xlsx")
sheet = book.active
dict = {}  # empty dictionary to store test data
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Sapushka"  # assigning value to the doc
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet["A3"].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestCase2":  # to scan specific data for specific test case below

        for j in range(2, sheet.max_column + 1):  # to get columns
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)
