import openpyxl


class HomePageData:
    test_formPage_userData = [
        {"Name": "John Cena", "Email": "TestVasyan@gmail.com", "Password": "Ololololo", "Gender": "Male"},
        {"Name": "Firefox", "Email": "2@gmail.com", "Password": "KEK", "Gender": "Female"},
        {"Name": "Rey Mysterio", "Email": "wwwe@wwe.com", "Password": "OLN",
         "Gender": "Male"}]

    @staticmethod  # allows usage of this method without creation of the class object
    def getTestData(test_case_name):
        dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\Alex\\PycharmProjects\\pythonAQAdraft\\PythonSelFramework\\TestData\\DemoFile.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i,
                          column=1).value == test_case_name:  # to scan specific data for specific test case below

                for j in range(2, sheet.max_column + 1):  # to get columns
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]  # we return dictionary as list [] since
